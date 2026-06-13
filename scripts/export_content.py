from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "portfolio_content_master.xlsx"
JSON_PATH = ROOT / "assets" / "content.json"
JS_PATH = ROOT / "assets" / "content.generated.js"
LANGS = ("en", "ja", "zh")


def rows_as_dicts(ws):
    header = [cell.value for cell in ws[1]]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        item = {}
        for key, value in zip(header, row):
            if key:
                item[key] = value
        items.append(item)
    return items


def clean(value):
    return "" if value is None else value


def copy_langs(prefix: str, row: dict, fallback_to_en: bool = True):
    values = {}
    en_value = clean(row.get(f"{prefix}_en"))
    for lang in LANGS:
        cell = clean(row.get(f"{prefix}_{lang}"))
        values[lang] = cell or (en_value if fallback_to_en else "")
    return values


def cell_langs(row: dict, lang_cols: dict[str, str], fallback_key: str | None = None):
    values = {}
    fallback = clean(row.get(fallback_key)) if fallback_key else ""
    for lang, col in lang_cols.items():
        cell = clean(row.get(col))
        values[lang] = cell or fallback
    return values


def build_content():
    wb = load_workbook(WORKBOOK_PATH, data_only=False)

    projects_rows = rows_as_dicts(wb["Projects"])
    project_text_rows = rows_as_dicts(wb["Project_Texts"])
    project_image_rows = rows_as_dicts(wb["Project_Images"])
    site_copy_rows = rows_as_dicts(wb["Site_Copy"])
    tag_rows = rows_as_dicts(wb["Tags"])
    about_content_rows = rows_as_dicts(wb["About_Content"])
    about_timeline_rows = rows_as_dicts(wb["About_Timeline"])
    contact_rows = rows_as_dicts(wb["Contact"])

    site_copy = defaultdict(dict)
    for row in site_copy_rows:
        site_copy[row["section"]][row["key"]] = {
            lang: clean(row.get(lang)) for lang in LANGS
        } | {"notes": clean(row.get("notes"))}

    texts_by_project = defaultdict(dict)
    for row in project_text_rows:
        project_id = row["project_id"]
        lang = row["lang"]
        texts_by_project[project_id][lang] = {
            "cardKicker": clean(row.get("card_kicker")),
            "cardTitle": clean(row.get("card_title")),
            "cardDescription": clean(row.get("card_description")),
            "detailTitle": clean(row.get("detail_title")),
            "detailIntro": clean(row.get("detail_intro")),
            "detailBody1": clean(row.get("detail_body_1")),
            "detailBody2": clean(row.get("detail_body_2")),
            "detailBody3": clean(row.get("detail_body_3")),
            "metaYear": clean(row.get("meta_year")),
            "metaStatus": clean(row.get("meta_status")),
            "metaLocation": clean(row.get("meta_location")),
            "metaRole": clean(row.get("meta_role")),
            "ctaLabel": clean(row.get("cta_label")),
            "translationStatus": clean(row.get("translation_status")),
            "lastUpdated": clean(row.get("last_updated")),
            "notes": clean(row.get("notes")),
        }

    images_by_project = defaultdict(lambda: defaultdict(list))
    for row in project_image_rows:
        image_order = row.get("image_order")
        images_by_project[row["project_id"]][row["image_role"]].append(
            {
                "imageOrder": int(image_order) if image_order not in (None, "") else None,
                "displayVariant": clean(row.get("display_variant")),
                "path": clean(row.get("image_path")),
                "alt": {
                    lang: clean(row.get(f"alt_{lang}")) or clean(row.get("alt_en"))
                    for lang in LANGS
                },
                "caption": {
                    lang: clean(row.get(f"caption_{lang}"))
                    or clean(row.get("caption_en"))
                    or clean(row.get(f"alt_{lang}"))
                    or clean(row.get("alt_en"))
                    for lang in LANGS
                },
                "credit": clean(row.get("credit")),
                "notes": clean(row.get("notes")),
            }
        )
    for project_images in images_by_project.values():
        for role_images in project_images.values():
            role_images.sort(
                key=lambda item: (
                    item.get("imageOrder") is None,
                    item.get("imageOrder") or 0,
                    item.get("path") or "",
                )
            )

    tags_by_project = defaultdict(list)
    for row in tag_rows:
        tags_by_project[row["project_id"]].append(
            {
                "displayOrder": row.get("display_order"),
                "label": {
                    "en": clean(row.get("tag_en")),
                    "ja": clean(row.get("tag_ja")) or clean(row.get("tag_en")),
                    "zh": clean(row.get("tag_zh")) or clean(row.get("tag_en")),
                },
            }
        )
    for values in tags_by_project.values():
        values.sort(key=lambda item: item.get("displayOrder") or 0)

    about_content = {
        row["key"]: {lang: clean(row.get(lang)) for lang in LANGS}
        for row in about_content_rows
    }
    about_timeline = []
    for row in sorted(about_timeline_rows, key=lambda item: item.get("sort_order") or 0):
        about_timeline.append(
            {
                "id": row["entry_id"],
                "sortOrder": row.get("sort_order") or 0,
                "year": clean(row.get("year")),
                "title": copy_langs("title", row),
                "location": copy_langs("location", row),
                "notes": clean(row.get("notes")),
            }
        )

    contacts = []
    for row in contact_rows:
        contacts.append(
            {
                "id": row["item_id"],
                "label": copy_langs("label", row),
                "value": clean(row.get("value")),
                "href": clean(row.get("href")),
                "notes": clean(row.get("notes")),
            }
        )

    projects = []
    for row in sorted(projects_rows, key=lambda item: item.get("sort_order") or 0):
        project_id = row["project_id"]
        cover = images_by_project[project_id].get("cover", [])
        gallery = images_by_project[project_id].get("gallery", [])
        projects.append(
            {
                "id": project_id,
                "slug": clean(row.get("slug")),
                "sortOrder": row.get("sort_order") or 0,
                "featured": bool(row.get("featured")),
                "category": clean(row.get("category")),
                "status": {
                    "en": clean(row.get("status_en")),
                    "ja": clean(row.get("status_ja")) or clean(row.get("status_en")),
                    "zh": clean(row.get("status_zh")) or clean(row.get("status_en")),
                },
                "yearDisplay": clean(row.get("year_display")),
                "location": {
                    "en": clean(row.get("location_en")),
                    "ja": clean(row.get("location_ja")) or clean(row.get("location_en")),
                    "zh": clean(row.get("location_zh")) or clean(row.get("location_en")),
                },
                "role": {
                    "en": clean(row.get("role_en")),
                    "ja": clean(row.get("role_ja")) or clean(row.get("role_en")),
                    "zh": clean(row.get("role_zh")) or clean(row.get("role_en")),
                },
                "collaboratingFirm": clean(row.get("collaborating_firm")),
                "sourceUrl": clean(row.get("source_url")),
                "notes": clean(row.get("notes")),
                "tags": tags_by_project.get(project_id, []),
                "images": {
                    "cover": cover[0] if cover else None,
                    "gallery": gallery,
                },
                "translations": {
                    lang: texts_by_project.get(project_id, {}).get(lang, {})
                    for lang in LANGS
                },
            }
        )

    content = {
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceWorkbook": WORKBOOK_PATH.name,
            "languages": list(LANGS),
        },
        "site": {
            "copy": site_copy,
            "about": {
                "summary": about_content.get("summary", {}),
                "born": {
                    "year": about_content.get("born_year", {}),
                    "label": about_content.get("born_label", {}),
                },
                "present": {
                    "year": about_content.get("present_year", {}),
                },
                "future": {
                    "note": about_content.get("future_note", {}),
                },
                "timeline": about_timeline,
            },
            "contact": {
                "items": contacts,
            },
        },
        "projects": projects,
    }

    return content


def main():
    content = build_content()
    JSON_PATH.write_text(
        json.dumps(content, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    JS_PATH.write_text(
        "window.__SITE_CONTENT_JSON__ = "
        + json.dumps(content, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"wrote {JSON_PATH}")
    print(f"wrote {JS_PATH}")


if __name__ == "__main__":
    main()
